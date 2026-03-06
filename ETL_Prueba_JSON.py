# -*- coding: utf-8 -*-

import json
import os
import re
import sys
import unicodedata
from datetime import datetime
from io import BytesIO
from typing import Dict, Optional, List

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU


# =========================================================
# CONFIG
# =========================================================

AUTO_ROW_HEIGHT = False
AUTO_ROW_MIN_HEIGHT = 34
AUTO_ROW_MAX_HEIGHT = 260
AUTO_LINE_HEIGHT = 23

EXCEL_WEB_MODE = True
EXCEL_WEB_MIN_ROW_HEIGHT = 38


# =========================================================
# Heurísticas
# =========================================================

TITLECASE_HINTS = (
    "estado", "municip", "colonia", "localizacion", "localización",
    "direccion", "dirección", "compania", "compañia", "nombre",
    "paterno", "materno", "puesto", "del_cd_mun", "cd_mun", "oficina",
    "region", "región"
)

SENTENCECASE_HINTS = (
    "proyecto", "descripcion", "descripción", "observaciones", "acabados"
)

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
CODE_RE  = re.compile(r"^[A-Z]{1,6}\d{2,}$")

CITY_FIXES = {
    "ciudad de mexico": "Ciudad de México",
    "estado de mexico": "Estado de México",
    "nuevo leon": "Nuevo León",
    "san luis potosi": "San Luis Potosí",
    "michoacan": "Michoacán",
    "queretaro": "Querétaro",
    "yucatan": "Yucatán",
    "aguascalientes": "Aguascalientes",
    "baja california": "Baja California",
    "baja california sur": "Baja California Sur",
    "campeche": "Campeche",
    "chihuahua": "Chihuahua",
    "coahuila": "Coahuila",
    "colima": "Colima",
    "durango": "Durango",
    "guanajuato": "Guanajuato",
    "guerrero": "Guerrero",
    "hidalgo": "Hidalgo",
    "jalisco": "Jalisco",
    "morelos": "Morelos",
    "nayarit": "Nayarit",
    "oaxaca": "Oaxaca",
    "puebla": "Puebla",
    "quintana roo": "Quintana Roo",
    "sinaloa": "Sinaloa",
    "sonora": "Sonora",
    "tabasco": "Tabasco",
    "tamaulipas": "Tamaulipas",
    "tlaxcala": "Tlaxcala",
    "veracruz": "Veracruz",
    "zacatecas": "Zacatecas",
}


# =========================================================
# CLÁSICO: Renombres / Clasificación / Orden
# =========================================================

RENAME_HEADERS_CLASICO = {
    "Inversion": "Inversión (Pesos Mexicanos)",
    "Sup_Construida": "Sup Construida (m²)",
    "Sup_Urbanizada": "Sup Urbanizada (m²)",
    "Internet": "Pagina Web",
}

CLASICO_OUTPUT_ORDER = [
    "Clave_Proyecto",
    "Fecha_Publicacion", "Fecha_publicacion",
    "Tipo_Proyecto",
    "Proyecto",
    "Etapa",
    "Tipo_Desarrollo",

    "Región", "Region",
    "Estado_Proyecto",
    "Del_Cd_Mun_Proyecto",
    "Localizacion1",
    "C.P.", "C.P",

    "Inversion",
    "Fecha_Inicio",
    "Fecha_Terminacion",
    "Genero",
    "Subgenero",
    "Tipo_Obra",
    "Sector",
    "Sup_Construida",
    "Sup_Urbanizada",
    "Numero_Unidades",
    "Num_Niveles",
    "Descripcion",

    "Rol_Compania",
    "Compania",
    "Direccion_Compania",
    "Colonia",
    "C.P.Compania",
    "Del_Cd_Mun_Compania",
    "Estado_Compania",
    "Telefono1",
    "Telefono2",
    "Telefono3",
    "Internet",

    "Titulo_1",
    "Contacto_1",
    "Puesto_1",
    "Extension_1",
    "Email_1",

    "Titulo_2",
    "Contacto_2",
    "Puesto_2",
    "Extension_2",
    "Email_2",

    "Titulo_3",
    "Contacto_3",
    "Puesto_3",
    "Extension_3",
    "Email_3",

    "Acabados",
    "Observaciones",
    "Descripcion_Extra",
]

_CLASSICO_GROUP_BY_ORIG_HEADER: Dict[str, str] = {
    "Clave_Proyecto": "Información General",
    "Fecha_publicacion": "Información General",
    "Fecha_Publicacion": "Información General",
    "Tipo_Proyecto": "Información General",
    "Proyecto": "Información General",
    "Etapa": "Información General",
    "Tipo_Desarrollo": "Información General",

    "Region": "Ubicación del Proyecto",
    "Región": "Ubicación del Proyecto",
    "Estado_Proyecto": "Ubicación del Proyecto",
    "Del_Cd_Mun_Proyecto": "Ubicación del Proyecto",
    "Localizacion1": "Ubicación del Proyecto",
    "C.P.": "Ubicación del Proyecto",
    "C.P": "Ubicación del Proyecto",

    "Inversion": "Caracteristicas del Proyecto",
    "Fecha_Inicio": "Caracteristicas del Proyecto",
    "Fecha_Terminacion": "Caracteristicas del Proyecto",
    "Genero": "Caracteristicas del Proyecto",
    "Subgenero": "Caracteristicas del Proyecto",
    "Tipo_Obra": "Caracteristicas del Proyecto",
    "Sector": "Caracteristicas del Proyecto",
    "Sup_Construida": "Caracteristicas del Proyecto",
    "Sup_Urbanizada": "Caracteristicas del Proyecto",
    "Numero_Unidades": "Caracteristicas del Proyecto",
    "Num_Niveles": "Caracteristicas del Proyecto",
    "Descripcion": "Caracteristicas del Proyecto",

    "Rol_Compania": "Datos de la Compañia",
    "Compania": "Datos de la Compañia",
    "Direccion_Compania": "Datos de la Compañia",
    "Colonia": "Datos de la Compañia",
    "C.P.Compania": "Datos de la Compañia",
    "Del_Cd_Mun_Compania": "Datos de la Compañia",
    "Estado_Compania": "Datos de la Compañia",
    "Telefono1": "Datos de la Compañia",
    "Telefono2": "Datos de la Compañia",
    "Telefono3": "Datos de la Compañia",
    "Internet": "Datos de la Compañia",

    "Titulo_1": "Participantes y Contactos",
    "Contacto_1": "Participantes y Contactos",
    "Puesto_1": "Participantes y Contactos",
    "Extension_1": "Participantes y Contactos",
    "Email_1": "Participantes y Contactos",

    "Titulo_2": "Participantes y Contactos",
    "Contacto_2": "Participantes y Contactos",
    "Puesto_2": "Participantes y Contactos",
    "Extension_2": "Participantes y Contactos",
    "Email_2": "Participantes y Contactos",

    "Titulo_3": "Participantes y Contactos",
    "Contacto_3": "Participantes y Contactos",
    "Puesto_3": "Participantes y Contactos",
    "Extension_3": "Participantes y Contactos",
    "Email_3": "Participantes y Contactos",

    "Acabados": "Detalles Adicionales",
    "Observaciones": "Detalles Adicionales",
    "Descripcion_Extra": "Detalles Adicionales",
}

COLOR_ORANGE = "ED7D31"
COLOR_GRAY = "E7E6E6"

_GROUP_FILLS = {
    "Información General": PatternFill("solid", COLOR_ORANGE),
    "Ubicación del Proyecto": PatternFill("solid", COLOR_GRAY),
    "Caracteristicas del Proyecto": PatternFill("solid", COLOR_ORANGE),
    "Datos de la Compañia": PatternFill("solid", COLOR_GRAY),
    "Participantes y Contactos": PatternFill("solid", COLOR_ORANGE),
    "Detalles Adicionales": PatternFill("solid", COLOR_GRAY),
    "Otros": PatternFill("solid", COLOR_GRAY),
}

_GROUP_FONT = Font(name="Poppins", size=11, bold=True, color="000000")
_GROUP_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# =========================================================
# ✅ SOLUCIÓN DE CARACTERES: mojibake repair + unescape
# =========================================================

_REPL_CHAR = "�"

def _score_spanish(s: str) -> int:
    if not isinstance(s, str) or not s:
        return -10
    good = "áéíóúüñÁÉÍÓÚÜÑ"
    bad  = "ÃÂ�µ¥¢¤ðÐþÞ¿¡"
    score = 0
    score += sum(2 for ch in s if ch in good)
    score -= sum(2 for ch in s if ch in bad)
    score -= sum(3 for ch in s if ord(ch) < 32 and ch not in "\n\t\r")
    return score

def _try_redecode(s: str, from_enc: str, to_enc: str) -> Optional[str]:
    try:
        b = s.encode(from_enc)
        return b.decode(to_enc)
    except Exception:
        return None

def _repair_mojibake(s: str) -> str:
    """
    Repara mojibake común ES:
    - UTF-8 leído como latin1/cp1252 => 'MÃ©xico'
    - CP850/CP437 vs latin1 => 'Ni¥os', 'µREA'
    - ✅ NUEVO: C1 controls (U+0080–U+009F) típicos cuando CP850 se interpretó como texto “plano”
      Ej: 'inter\x90s' donde 0x90 (CP850) = 'É'
    """
    if not isinstance(s, str) or not s:
        return s

    # Detecta C1 controls (U+0080..U+009F)
    has_c1 = any(0x80 <= ord(ch) <= 0x9F for ch in s)

    symptoms = ("Ã", "Â", "�", "µ", "¥", "¤", "¢", "à", "¨")
    if (not has_c1) and (not any(x in s for x in symptoms)):
        return s

    candidates = [s]

    def try_redecode_latin1_to(target_enc: str):
        try:
            b = s.encode("latin1", errors="strict")   # latin1 preserva 0x00..0xFF tal cual
            return b.decode(target_enc, errors="strict")
        except Exception:
            return None

    def try_redecode(from_enc: str, to_enc: str):
        try:
            b = s.encode(from_enc)
            return b.decode(to_enc)
        except Exception:
            return None

    # Caso clásico: UTF-8 mal leído como Latin1/CP1252 => 'MÃ©xico'
    for enc in ("latin1", "cp1252"):
        cand = try_redecode(enc, "utf-8")
        if cand:
            candidates.append(cand)

    # Caso CP850/CP437 mal tratado como texto => aparecen ¥, µ, o C1 controls (\x90 etc)
    for target in ("cp850", "cp437"):
        cand = try_redecode_latin1_to(target)
        if cand:
            candidates.append(cand)

    # Tu ruta anterior (cp1252 <-> cp850/cp437)
    for target in ("cp850", "cp437"):
        cand = try_redecode("cp1252", target)
        if cand:
            candidates.append(cand)

    for src in ("cp850", "cp437"):
        cand = try_redecode(src, "cp1252")
        if cand:
            candidates.append(cand)

    # Scoring “español” (reusa tu _score_spanish si ya existe)
    def score_spanish(x: str) -> int:
        good = "áéíóúüñÁÉÍÓÚÜÑ"
        bad  = "ÃÂ�µ¥¢¤ðÐþÞ¿¡"
        sc = 0
        sc += sum(2 for ch in x if ch in good)
        sc -= sum(2 for ch in x if ch in bad)
        # ✅ penaliza C1 controls
        sc -= sum(4 for ch in x if 0x80 <= ord(ch) <= 0x9F)
        return sc

    best = max(candidates, key=score_spanish)

    # Si aún quedan C1 controls, al menos elimínalos para que no “corten” el texto
    best = "".join(ch for ch in best if not (0x80 <= ord(ch) <= 0x9F))

    return best
def _unescape_quotes_backslashes(s: str) -> str:
    if not isinstance(s, str) or "\\" not in s:
        return s
    s = s.replace(r'\\"', '"')
    s = s.replace(r"\"", '"')
    s = s.replace(r"\n", "\n").replace(r"\t", "\t")
    return s

def _repair_all_strings_df(df: pd.DataFrame) -> pd.DataFrame:
    def fix(v):
        if isinstance(v, str):
            v = _repair_mojibake(v)
            v = _unescape_quotes_backslashes(v)
        return v

    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].map(fix)
    return df


# =========================================================
# Helpers texto
# =========================================================

def _norm_noaccents_lower(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _strip_wrapping_quotes(s: str) -> str:
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        return s[1:-1]
    return s

def _sentence_case_spanish(s: str) -> str:
    s = s.lower()
    out = []
    cap_next = True
    for ch in s:
        if cap_next and ch.isalpha():
            out.append(ch.upper())
            cap_next = False
        else:
            out.append(ch)
        if ch in ".!?":
            cap_next = True
        if ch == "\n":
            cap_next = True
    return "".join(out)

def _title_case_spanish(s: str) -> str:
    lower_words = {"de","del","la","las","el","los","y","e","o","u","a","al","en","con","por","para","sin"}
    parts = s.split()
    out = []
    for i, w in enumerate(parts):
        wl = w.lower()
        if not any(ch.isalpha() for ch in w):
            out.append(w); continue
        if i > 0 and wl in lower_words:
            out.append(wl)
        else:
            out.append(wl[:1].upper() + wl[1:])
    return " ".join(out)

def _choose_case_strategy(col_name: str) -> str:
    c = (col_name or "").lower()
    if any(h in c for h in SENTENCECASE_HINTS):
        return "sentence"
    if any(h in c for h in TITLECASE_HINTS):
        return "title"
    return "none"

def _uppercase_inside_quotes(text: str) -> str:
    if not isinstance(text, str) or '"' not in text:
        return text
    return re.sub(r'"([^"]+)"', lambda m: '"' + m.group(1).upper() + '"', text)

def _fix_shouty_caps_mixed(text: str) -> str:
    if not isinstance(text, str):
        return text
    keep_upper = {"C.P", "C.P.", "ID", "PDF", "MX", "M2", "M²", "SAT", "IMEI", "NSS"}

    def fix_word(w: str) -> str:
        letters = [ch for ch in w if ch.isalpha()]
        if not letters:
            return w
        if w.replace(":", "").replace(".", "").upper() in {k.replace(".", "") for k in keep_upper}:
            return w
        if len([ch for ch in w if ch.isalpha()]) >= 3 and all(ch.isupper() for ch in letters):
            low = w.lower()
            return low[:1].upper() + low[1:]
        return w

    tokens = re.split(r'(\s+)', text)
    out = []
    for t in tokens:
        if t.isspace() or t == "":
            out.append(t); continue
        subtoks = re.split(r'([,:;.\-()])', t)
        fixed = []
        for st in subtoks:
            if st in {",", ":", ";", ".", "-", "(", ")"} or st == "":
                fixed.append(st)
            else:
                fixed.append(fix_word(st))
        out.append("".join(fixed))
    return "".join(out)

def _normalize_free_text(col: str, s: str, force_upper: bool = False) -> str:
    if not isinstance(s, str):
        return s
    s = _fix_shouty_caps_mixed(s)
    if force_upper:
        return s.upper()

    col_l = (col or "").lower()
    if col_l == "localizacion1":
        return _title_case_spanish(s)

    s = _sentence_case_spanish(s)
    if col_l == "proyecto":
        s = _uppercase_inside_quotes(s)
    return s

def _fix_clasico_observaciones_codes(s: str) -> str:
    if not isinstance(s, str):
        return s
    return re.sub(r"\b([oOpP])([cCpP])(\d{2,})\b",
                  lambda m: (m.group(1) + m.group(2)).upper() + m.group(3), s)

def _smart_text_format(v, col_name: str):
    if not isinstance(v, str):
        return v

    s = v.strip()
    s = _strip_wrapping_quotes(s)
    if not s:
        return None

    k = _norm_noaccents_lower(s)
    if k in CITY_FIXES:
        return CITY_FIXES[k]

    if EMAIL_RE.match(s):
        return s.lower()

    if CODE_RE.match(s):
        return s

    if " " not in s and any(ch.isdigit() for ch in s) and any(ch.isalpha() for ch in s):
        return s

    letters = [ch for ch in s if ch.isalpha()]
    all_caps = bool(letters) and all(ch.isupper() for ch in letters)
    if all_caps:
        strat = _choose_case_strategy(col_name)
        if strat == "sentence":
            return _sentence_case_spanish(s)
        if strat == "title":
            return _title_case_spanish(s)
        return _title_case_spanish(s)

    return s


# =========================================================
# Helpers DF / Excel
# =========================================================

def _export_headers_with_spaces(cols):
    return [str(c).replace("_", " ").strip() for c in cols]

def _compute_widths_from_df(df: pd.DataFrame, padding: int = 4, max_width: int = 60) -> dict:
    if df.empty:
        return {}
    sample = df if len(df) <= 2000 else df.head(2000)
    widths = {}
    for col in sample.columns:
        max_len = len(str(col))
        for v in sample[col]:
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        widths[col] = min(max_len + padding, max_width)
    return widths

def _norm_colkey(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace(" ", "_").replace("-", "_")
    s = re.sub(r"_+", "_", s)
    return s

def _apply_width_overrides(ws, df_export: pd.DataFrame):
    TARGET_WIDTHS = {
        "email_1": 42,
        "email_2": 42,
        "email_3": 55,
        "proyecto": 95,
        "descripcion": 120,
        "descripcionextra": 120,
        "observaciones": 95,
        "descripcion_extra": 120,
        "localizacion": 85,
        "localizacion1": 85,
        "localizacion_del_proyecto": 95,
    }

    for idx, col_name in enumerate(df_export.columns, start=1):
        k = _norm_colkey(col_name)
        if k == "email3": k = "email_3"
        if k == "email2": k = "email_2"
        if k == "email1": k = "email_1"
        if k in ("descripcion_extra_del_proyecto",): k = "descripcion_extra"
        if k in TARGET_WIDTHS:
            ws.column_dimensions[get_column_letter(idx)].width = TARGET_WIDTHS[k]

def _resolve_resource_path(filename: str) -> Optional[str]:
    base_path = getattr(sys, "_MEIPASS", None)
    if base_path:
        candidate = os.path.join(base_path, filename)
        if os.path.exists(candidate):
            return candidate
    candidate = os.path.join(os.path.dirname(__file__), filename)
    if os.path.exists(candidate):
        return candidate
    candidate = os.path.join(os.getcwd(), filename)
    if os.path.exists(candidate):
        return candidate
    return None

def _is_clasico(tipo_reporte: str) -> bool:
    return str(tipo_reporte).strip().lower() == "clasico"

def _apply_explicit_order_clasico(df: pd.DataFrame) -> pd.DataFrame:
    ordered = [c for c in CLASICO_OUTPUT_ORDER if c in df.columns]
    rest = [c for c in df.columns if c not in ordered]
    return df[ordered + rest]

def _drop_otros_columns_for_clasico(df: pd.DataFrame) -> pd.DataFrame:
    keep_cols = [c for c in df.columns if _CLASSICO_GROUP_BY_ORIG_HEADER.get(c, "Otros") != "Otros"]
    return df[keep_cols]

def _drop_localizacion2(df: pd.DataFrame) -> pd.DataFrame:
    def key(x: str) -> str:
        return _norm_noaccents_lower(x).replace(" ", "_").replace("__", "_")
    drop = [c for c in df.columns if key(c) in ("localizacion2", "localizacion_2")]
    if drop:
        df = df.drop(columns=drop)
    return df

def _build_contacto(df: pd.DataFrame, n: int) -> pd.Series:
    nom = f"Nombre_{n}"
    pat = f"Paterno_{n}"
    mat = f"Materno_{n}"

    def join_row(row):
        parts = []
        for k in (nom, pat, mat):
            v = row.get(k, None)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            s = str(v).strip()
            if not s or s.lower() == "nan":
                continue
            parts.append(s)
        return " ".join(parts).strip() if parts else None

    return df.apply(join_row, axis=1)


# =========================================================
# Estilos / formatos
# =========================================================

def _apply_styles_excel_and_sheets(ws, header_row: int, first_data_row: int, nrows: int, ncols: int, orig_headers=None):
    header_font = Font(name="Poppins", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Poppins", size=11)

    header_fill = PatternFill("solid", COLOR_ORANGE)
    zebra_a = PatternFill("solid", "F2F2F2")
    zebra_b = PatternFill("solid", "FFFFFF")

    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    body_align_left = Alignment(wrap_text=True, vertical="center", horizontal="left")

    LEFT_ALIGN_COLUMNS = {
        "proyecto",
        "localizacion1",
        "descripcion",
        "acabados",
        "observaciones"
    }

    ws.freeze_panes = f"A{first_data_row}"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{nrows}"

    for col_idx, cell in enumerate(ws[header_row], start=1):

        cell.alignment = header_align

        if header_row == 3 and orig_headers:

            orig_header = orig_headers[col_idx - 1]
            group_name = _CLASSICO_GROUP_BY_ORIG_HEADER.get(orig_header, "Otros")
            fill = _GROUP_FILLS.get(group_name, _GROUP_FILLS["Otros"])

            cell.fill = fill

            # fuente dinámica según fondo
            if fill.start_color.rgb.endswith(COLOR_ORANGE):
                cell.font = Font(name="Poppins", size=11, bold=True, color="FFFFFF")
            else:
                cell.font = Font(name="Poppins", size=11, bold=True, color="000000")

        else:
            cell.fill = header_fill
            cell.font = header_font

    for r in range(first_data_row, nrows + 1):

        fill = zebra_a if (r % 2 == 0) else zebra_b

        for c in range(1, ncols + 1):

            cell = ws.cell(row=r, column=c)
            header_name = str(ws.cell(row=header_row, column=c).value)

            norm = header_name.lower().replace(" ", "_")

            cell.font = body_font
            cell.fill = fill

            if norm in LEFT_ALIGN_COLUMNS:
                cell.alignment = body_align_left
            else:
                cell.alignment = body_align_center


def _apply_section_borders(ws, header_row: int, first_data_row: int, nrows: int, ncols: int, orig_headers: List[str]):
    """
    Dibuja un borde vertical en el cambio de sección (entre grupos).
    No dibuja bordes entre columnas del mismo grupo.
    """
    if not orig_headers:
        return

    # Borde que usaremos para separar secciones
    section_side = Side(style="thin", color="000000")
    section_border = Border(left=section_side)

    # Determinar el grupo de cada columna
    groups = [_CLASSICO_GROUP_BY_ORIG_HEADER.get(h, "Otros") for h in orig_headers]

    # Columnas donde cambia el grupo
    section_starts = []
    for i in range(1, len(groups)):
        if groups[i] != groups[i - 1]:
            section_starts.append(i + 1)  # +1 porque Excel es 1-index

    # Aplicar borde a todas las filas visibles del reporte
    for col in section_starts:
        for row in range(header_row - 1, nrows + 1):
            cell = ws.cell(row=row, column=col)

            existing = cell.border

            cell.border = Border(
                left=section_side,
                right=existing.right,
                top=existing.top,
                bottom=existing.bottom
            )


def _apply_row_borders(ws, first_data_row: int, nrows: int, ncols: int):

    row_side = Side(style="thin", color="b0b0b0")

    for row in range(first_data_row, nrows + 1):
        for col in range(1, ncols + 1):

            cell = ws.cell(row=row, column=col)

            existing = cell.border

            cell.border = Border(
                left=existing.left,
                right=existing.right,
                top=existing.top,
                bottom=row_side
            )

def _apply_fixed_row_height(ws, first_data_row: int, nrows: int, height: int = 38):

    for r in range(first_data_row, nrows + 1):
        ws.row_dimensions[r].height = height

def _format_date_columns_no_time(ws, df_orig: pd.DataFrame, first_data_row: int):
    date_cols_idx = [i for i, col in enumerate(df_orig.columns, start=1) if "fecha" in str(col).lower()]
    if not date_cols_idx:
        return
    date_fmt = "yyyy-mm-dd"
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in range(first_data_row, ws.max_row + 1):
        for c in date_cols_idx:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, datetime):
                cell.number_format = date_fmt
                cell.alignment = left_align

def _format_numeric_columns(ws, df_orig: pd.DataFrame, first_data_row: int):
    formats = {}
    for i, col in enumerate(df_orig.columns, start=1):
        cname = _norm_noaccents_lower(str(col)).replace(" ", "_")

        if "inversion" in cname:
            formats[i] = '$#,##0'
        elif col in ("Sup_Construida", "Sup_Urbanizada"):
            formats[i] = '#,##0'
        elif "sup_" in cname or "superficie" in cname:
            formats[i] = '#,##0'
        elif cname in ("numero_unidades", "num_niveles"):
            formats[i] = '#,##0'
        elif cname in ("latitud", "longitud", "latitude", "longitude", "lat", "lng"):
            formats[i] = '0.000000'
        elif cname in ("dia_publicado", "mes_publicado", "ano_publicado", "anio_publicado",
                       "dia_inicio", "mes_inicio", "ano_inicio", "anio_inicio"):
            formats[i] = '0'
        else:
            if col in df_orig.columns and pd.api.types.is_numeric_dtype(df_orig[col]):
                formats[i] = '#,##0'

    for r in range(first_data_row, ws.max_row + 1):
        for cidx, fmt in formats.items():
            cell = ws.cell(row=r, column=cidx)
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = fmt


# =========================================================
# Branding / Clasificación
# =========================================================

def _apply_branding_row(
    ws,
    ncols: int,
    empresa: str,
    usuario: str,
    fecha_descarga: str,
    report_label: str,
    logo_filename: str = "logo_bimsa.jpg",
    tipo_fecha: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    logo_path: Optional[str] = None,
):
    # ws.row_dimensions[1].height = 60

    white = PatternFill("solid", "FFFFFF")
    bottom_side = Side(style="thin", color="000000")

    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = white

        existing = cell.border

        cell.border = Border(
            left=existing.left,
            right=existing.right,
            top=existing.top,
            bottom=bottom_side
        )

    ws.column_dimensions["A"].width = 30

    final_logo_path = logo_path if (logo_path and os.path.exists(logo_path)) else _resolve_resource_path(logo_filename)

    if final_logo_path:
        img = XLImage(final_logo_path)
        img.width = 200
        img.height = 50

        marker = AnchorMarker(
            col=0,
            colOff=0,
            row=0,
            rowOff=120000  # ← centra verticalmente
        )

        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(
                pixels_to_EMU(img.width),
                pixels_to_EMU(img.height)
            )
        )

        ws.add_image(img)

    else:
        ws["A1"].value = "BimsaReports"
        ws["A1"].font = Font(name="Poppins", size=18, bold=True, color="666666")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)

    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45

    info_cell = ws.cell(1, 2)
    info_lines = [
        f"Empresa: {empresa}",
        f"Usuario: {usuario}",
        f"Fecha de descarga: {fecha_descarga}",
    ]

    info_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    if tipo_fecha and fecha_inicio and fecha_fin:
        info_lines.append(
            f'Descarga por {tipo_fecha} del {fecha_inicio} al {fecha_fin}'
        )

    info_cell.value = "\n".join(info_lines)
    info_cell.font = Font(name="Poppins", size=11, bold=False, color="000000")
    info_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    right_col = ncols
    ws.merge_cells(start_row=1, start_column=max(1, right_col - 1), end_row=1, end_column=right_col)
    tag_cell = ws.cell(1, max(1, right_col - 1))
    tag_cell.value = report_label
    tag_cell.font = Font(name="Poppins", size=14, bold=True, color="FFFFFF")
    tag_cell.fill = PatternFill("solid", "ED7D31")
    tag_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for c in range(1, ncols + 1):
        ws.cell(1, c).fill = white

def _apply_classification_row(ws, header_row: int, ncols: int, orig_headers: List[str]):
    cls_row = header_row - 1
    ws.row_dimensions[cls_row].height = 22

    groups = [_CLASSICO_GROUP_BY_ORIG_HEADER.get(h, "Otros") for h in orig_headers]

    c = 1
    while c <= ncols:
        g = groups[c - 1]
        start = c
        end = c
        while end + 1 <= ncols and groups[end] == g:
            end += 1

        ws.merge_cells(start_row=cls_row, start_column=start, end_row=cls_row, end_column=end)
        cell = ws.cell(cls_row, start)
        cell.value = g
        fill = _GROUP_FILLS.get(g, _GROUP_FILLS["Otros"])
        cell.fill = fill

        # fuente dinámica según color
        if fill.start_color.rgb.endswith(COLOR_ORANGE):
            cell.font = Font(name="Poppins", size=11, bold=True, color="FFFFFF")
        else:
            cell.font = Font(name="Poppins", size=11, bold=True, color="000000")

        cell.alignment = _GROUP_ALIGN

        for cc in range(start, end + 1):
            ws.cell(cls_row, cc).fill = fill

        c = end + 1


# =========================================================
# Auto-height (wrap por palabras)
# =========================================================

def _safe_col_width(ws, col_letter: str) -> float:
    w = ws.column_dimensions[col_letter].width
    return float(w) if (w is not None and w > 0) else 12.0

def _estimate_wrapped_lines(text: str, col_width_chars: float) -> int:
    if not text:
        return 1
    SHEETS_WIDTH_FACTOR = 0.88
    effective = max(8, int(col_width_chars * SHEETS_WIDTH_FACTOR) - 1)

    total_lines = 0
    for raw_line in str(text).split("\n"):
        raw_line = raw_line.strip()
        if not raw_line:
            total_lines += 1
            continue

        words = raw_line.split()
        line_len = 0
        lines = 1

        for w in words:
            wlen = len(w)
            if line_len == 0:
                line_len = wlen
            elif line_len + 1 + wlen <= effective:
                line_len += 1 + wlen
            else:
                lines += 1
                line_len = wlen

        total_lines += lines

    return max(1, total_lines)

def _apply_auto_row_heights(ws, first_data_row: int, min_h: int, max_h: int, line_h: int):
    for r in range(first_data_row, ws.max_row + 1):
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            col_letter = get_column_letter(c)
            col_w = _safe_col_width(ws, col_letter)
            lines = _estimate_wrapped_lines(cell.value, col_w)
            max_lines = max(max_lines, lines)

        ws.row_dimensions[r].height = max(min_h, min(max_h, int(max_lines * line_h)))


# =========================================================
# ETL principal
# =========================================================

def ETL_BIMSA(
    ruta_json: str,
    tipo_reporte: str,
    return_mode: str = "file",
    carpeta_excel: str = ".",
    empresa: str = "",
    usuario: str = "",
    tipo_fecha: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    report_label: Optional[str] = None,
    logo_path: Optional[str] = None,
):
    print("[BIMSA_ETL] Iniciando ETL BIMSA...")

    tipo_upper = str(tipo_reporte).strip().upper()
    es_clasico = _is_clasico(tipo_reporte)

    now = datetime.now()
    fecha_descarga = now.strftime("%d-%m-%Y")
    nombre_excel = f"BIMSA_{tipo_upper}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    if report_label is None:
        report_label = tipo_upper.lower()

    with open(ruta_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list) or not data:
        raise ValueError("El JSON debe ser una lista con al menos un registro")

    df = pd.DataFrame(data)

    # Nunca Localizacion2
    df = _drop_localizacion2(df)

    # ✅ FIX GLOBAL DE CARACTERES (a TODO texto)
    df = _repair_all_strings_df(df)

    # Columnas catálogo (no deben procesarse)
    CATALOGO_COLUMNS = {
        "tipo_proyecto",
        "etapa",
        "tipo_desarrollo",
        "region",
        "estado_proyecto",
        "estado",
        "genero",
        "subgenero",
        "tipo_obra",
        "sector",
        "rol_compania",
        "estado_compania",
        "puesto",
        "puesto_1",
        "puesto_2",
        "puesto_3",
    }

    for col in df.select_dtypes(include="object").columns:

        norm_col = _norm_colkey(col)

        if norm_col in CATALOGO_COLUMNS:
            continue

        df[col] = df[col].map(lambda v, c=col: _smart_text_format(v, c))

    # Reglas por tipo
    force_upper_proyecto_mapas = (tipo_upper == "MAPAS")

    ALIAS_NOMBRE_PROY = {"nombre_del_proyecto"}
    ALIAS_DESC_PROY   = {"descripcion_del_proyecto"}
    ALIAS_PROYECTO    = {"proyecto", "nombre"}
    ALIAS_LOCALIZ     = {"localizacion1", "localizacion", "localizacion_del_proyecto"}
    ALIAS_OBSERV      = {"observaciones"}
    ALIAS_DESC_EXTRA  = {"descripcion_extra", "descripcionextra", "descripcion_extra_del_proyecto"}

    for col in list(df.columns):
        nk2 = _norm_noaccents_lower(col).replace(" ", "_").replace("__", "_")

        if tipo_upper == "CONTACTOS":
            if nk2 in ALIAS_NOMBRE_PROY:
                df[col] = df[col].map(lambda x: _normalize_free_text("Proyecto", x, force_upper=True))
                continue
            if nk2 in ALIAS_DESC_PROY:
                df[col] = df[col].map(lambda x: _normalize_free_text("Proyecto", x, force_upper=False))
                df[col] = df[col].map(_uppercase_inside_quotes)
                continue

        if tipo_upper == "MAPAS" and nk2 in (ALIAS_NOMBRE_PROY | ALIAS_DESC_PROY | ALIAS_PROYECTO):
            df[col] = df[col].map(lambda x: _normalize_free_text("Proyecto", x, force_upper=True))
            continue

        if nk2 in (ALIAS_NOMBRE_PROY | ALIAS_DESC_PROY | ALIAS_PROYECTO):
            df[col] = df[col].map(lambda x: _normalize_free_text("Proyecto", x, force_upper=False))
            df[col] = df[col].map(_uppercase_inside_quotes)

        elif nk2 in ALIAS_LOCALIZ:
            df[col] = df[col].map(lambda x: _normalize_free_text("Localizacion1", x, force_upper=False))

        elif nk2 in ALIAS_OBSERV:
            df[col] = df[col].map(lambda x: _normalize_free_text("Observaciones", x, force_upper=False))
            if es_clasico:
                df[col] = df[col].map(_fix_clasico_observaciones_codes)

        elif nk2 in ALIAS_DESC_EXTRA:
            df[col] = df[col].map(lambda x: _normalize_free_text("Descripcion_Extra", x, force_upper=False))

    # Descripcion / Acabados
    for c in ("Descripcion", "Acabados"):
        if c in df.columns:
            df[c] = df[c].map(_fix_shouty_caps_mixed)
            df[c] = df[c].map(lambda x: _sentence_case_spanish(x) if isinstance(x, str) else x)

    # Fechas
    for col in df.columns:
        if "fecha" in str(col).lower():
            s = df[col]
            has_slash = False
            try:
                sample = s.dropna().astype(str).head(50)
                has_slash = any("/" in v for v in sample)
            except Exception:
                pass
            df[col] = pd.to_datetime(s, errors="coerce", dayfirst=has_slash)

    # Numéricos generales
    for col in df.columns:
        cname = str(col).lower()
        if "inversion" in cname:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        if "sup_" in cname or "superficie" in cname:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    for c in ("Numero_Unidades", "Num_Niveles"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # MAPAS: lat/long + dia/mes/año numéricos
    if tipo_upper == "MAPAS":
        def _k(name: str) -> str:
            return _norm_noaccents_lower(name).replace(" ", "_").replace("__", "_")
        for col in df.columns:
            kk = _k(col)
            if kk in ("latitud", "longitud", "latitude", "longitude", "lat", "lng"):
                df[col] = pd.to_numeric(df[col], errors="coerce")
            if kk in (
                "dia_publicado", "mes_publicado", "ano_publicado", "anio_publicado",
                "dia_inicio", "mes_inicio", "ano_inicio", "anio_inicio"
            ):
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

    # CLASICO: contactos + orden + drop otros
    if es_clasico:
        for n in (1, 2, 3):
            df[f"Contacto_{n}"] = _build_contacto(df, n)

        drop_cols = []
        for n in (1, 2, 3):
            drop_cols += [f"Nombre_{n}", f"Paterno_{n}", f"Materno_{n}"]
        df = df.drop(columns=[c for c in drop_cols if c in df.columns])

        df = _apply_explicit_order_clasico(df)
        df = _drop_otros_columns_for_clasico(df)

    # Export DF
    df_export = df.copy()
    if es_clasico:
        df_export = df_export.rename(columns=RENAME_HEADERS_CLASICO)
    df_export.columns = _export_headers_with_spaces(df_export.columns)

    # OUTPUT
    if return_mode == "bytes":
        output = BytesIO()
        writer_target = output
        ruta_final = None
    else:
        os.makedirs(carpeta_excel, exist_ok=True)
        ruta_final = f"{carpeta_excel.rstrip('/')}/{nombre_excel}"
        writer_target = ruta_final

    # Write Excel
    sheet_name = "Reporte"
    with pd.ExcelWriter(writer_target, engine="openpyxl") as writer:
        header_row = 3 if es_clasico else 2
        startrow = header_row - 1
        first_data_row = header_row + 1

        df_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
        ws = writer.sheets[sheet_name]

        ncols = ws.max_column
        nrows = ws.max_row

        if es_clasico:
            _apply_classification_row(ws, header_row=header_row, ncols=ncols, orig_headers=list(df.columns))

        widths = _compute_widths_from_df(df_export, padding=4, max_width=60)
        for idx, col_name in enumerate(df_export.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 12)

        for idx, col_name in enumerate(df_export.columns, start=1):
            if str(col_name).strip().lower() in ("descripcion", "descripción", "acabados"):
                ws.column_dimensions[get_column_letter(idx)].width = 115

        _apply_width_overrides(ws, df_export)

        _apply_styles_excel_and_sheets(
            ws,
            header_row,
            first_data_row,
            nrows,
            ncols,
            orig_headers=list(df.columns) if es_clasico else None
        )
        if es_clasico:
            _apply_section_borders(
                ws,
                header_row=header_row,
                first_data_row=first_data_row,
                nrows=nrows,
                ncols=ncols,
                orig_headers=list(df.columns),
            )
            _apply_row_borders(
                ws,
                first_data_row=first_data_row,
                nrows=nrows,
                ncols=ncols
            )

        ws.row_dimensions[header_row].height = 38


        _format_date_columns_no_time(ws, df, first_data_row)
        _format_numeric_columns(ws, df, first_data_row)

        _apply_branding_row(
            ws,
            ncols=ncols,
            empresa=empresa,
            usuario=usuario,
            fecha_descarga=fecha_descarga,
            report_label=report_label,
            tipo_fecha=tipo_fecha,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            logo_filename="logo_bimsa.jpg",
            logo_path=logo_path,
        )

        if AUTO_ROW_HEIGHT:
            min_h = AUTO_ROW_MIN_HEIGHT
            if EXCEL_WEB_MODE:
                min_h = max(min_h, EXCEL_WEB_MIN_ROW_HEIGHT)

            _apply_auto_row_heights(
                ws,
                first_data_row=first_data_row,
                min_h=min_h,
                max_h=AUTO_ROW_MAX_HEIGHT,
                line_h=AUTO_LINE_HEIGHT,
            )
        else:
            _apply_fixed_row_height(
                ws,
                first_data_row=first_data_row,
                nrows=nrows,
                height=60
            )


    print(f"[BIMSA_ETL] ETL terminado correctamente: {nombre_excel}")

    if return_mode == "bytes":
        output.seek(0)
        return nombre_excel, output.getvalue()

    return ruta_final